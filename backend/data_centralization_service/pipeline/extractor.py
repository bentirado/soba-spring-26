"""Extract stage — reads raw source files from dummy_data/."""

from __future__ import annotations

import datetime
import json
from pathlib import Path
from typing import Any

import openpyxl

from ..config import settings
from ..schema import ExtractedDataset


class DummyDataExtractor:
    """Extracts local source files from backend/dummy_data."""

    def __init__(
        self,
        dummy_data_dir: Path | None = None,
        volunteers_file: Path | None = None,
    ):
        self.dummy_data_dir = dummy_data_dir or settings.dummy_data_directory
        self.volunteers_file = volunteers_file or settings.mock_volunteers_file

    def extract(self) -> list[ExtractedDataset]:
        datasets: list[ExtractedDataset] = []

        if self.volunteers_file.exists():
            volunteers = self._load_json(self.volunteers_file)
            datasets.append(
                ExtractedDataset(
                    source_name="mock_volunteers",
                    source_type="json_records",
                    payload=volunteers,
                    metadata={
                        "path": str(self.volunteers_file),
                        "record_count": len(volunteers),
                    },
                )
            )

        for xlsx_path in sorted(self.dummy_data_dir.glob("*.xlsx")):
            sheets = self._load_excel(xlsx_path)
            datasets.append(
                ExtractedDataset(
                    source_name=xlsx_path.stem,
                    source_type="excel_file",
                    payload=sheets,
                    metadata={
                        "path": str(xlsx_path),
                        "sheet_names": list(sheets.keys()),
                        "total_rows": sum(len(rows) for rows in sheets.values()),
                    },
                )
            )

        return datasets

    def _load_json(self, path: Path) -> list[dict[str, Any]]:
        with path.open("r", encoding="utf-8") as handle:
            return json.load(handle)

    def _load_excel(self, path: Path) -> dict[str, list[dict[str, Any]]]:
        """Read all sheets from an Excel file into header-keyed row dicts."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                sheets[sheet_name] = []
                continue
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            records: list[dict[str, Any]] = []
            for row in rows[1:]:
                if all(cell is None for cell in row):
                    continue
                record = {}
                for key, value in zip(headers, row):
                    record[key] = _serialize_cell(value)
                records.append(record)
            sheets[sheet_name] = records
        wb.close()
        return sheets


def _serialize_cell(value: Any) -> Any:
    """Convert Excel cell values to JSON-safe types."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.isoformat()
    if isinstance(value, datetime.timedelta):
        total_seconds = int(value.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return value
